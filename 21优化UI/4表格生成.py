# -*- coding: utf-8 -*-
import os
import json
import requests
import tkinter as tk
from tkinter import filedialog, messagebox
from pathlib import Path
from openpyxl import Workbook

# ================= 路径 =================
CACHE_FILE = Path.cwd() / "config_excel.json"

# ================= 默认配置 =================
DEFAULT_CONFIG = {
    "phone": "授权码",
    "account_name": "账户名称",
    "image_dir": "",
    "shop_name": "TEMU户外2店",
    "spuid": "604448109735250",
    "output_excel": ""
}

# ================= 配置 =================
def load_config():
    if CACHE_FILE.exists():
        try:
            return {**DEFAULT_CONFIG, **json.loads(CACHE_FILE.read_text(encoding="utf-8"))}
        except:
            return DEFAULT_CONFIG
    else:
        save_config(DEFAULT_CONFIG)
        return DEFAULT_CONFIG

def save_config(data):
    CACHE_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

config = load_config()

# ================= 授权 =================
def check_auth(phone, account):
    url = "http://hk.ludada.vip:6688/check"
    try:
        res = requests.post(url, json={"phone": phone, "account_name": account}, timeout=10)
        result = res.json()
        return result.get("code") == 1, result.get("msg", "")
    except Exception as e:
        return False, str(e)

# ================= 主逻辑 =================
def process():
    image_folder = image_var.get().strip()
    output_excel = output_var.get().strip()

    config["phone"] = phone_var.get().strip()
    config["account_name"] = account_var.get().strip()
    config["image_dir"] = image_folder
    config["shop_name"] = shop_var.get().strip()
    config["spuid"] = spuid_var.get().strip()
    config["output_excel"] = output_excel

    save_config(config)

    # 授权验证
    ok, msg = check_auth(config["phone"], config["account_name"])
    if not ok:
        messagebox.showerror("授权失败", msg)
        return

    if not os.path.exists(image_folder):
        messagebox.showerror("错误", "图片文件夹不存在")
        return

    image_files = sorted([
        f for f in os.listdir(image_folder)
        if f.lower().endswith(('.png', '.jpg', '.jpeg'))
    ])

    if not image_files:
        messagebox.showwarning("提示", "没有图片")
        return

    # 默认输出路径
    if not output_excel:
        output_excel = os.path.join(image_folder, "上架表模板MX_已处理.xlsx")

    # 避免覆盖
    if os.path.exists(output_excel):
        base, ext = os.path.splitext(output_excel)
        output_excel = f"{base}_1{ext}"

    wb = Workbook()
    sheet = wb.active

    # 表头
    headers = [
        "素材本机地址", "素材图", "详情图list", "主图list",
        "SKU货号编码list", "本地图片包地址下文件列表",
        "店铺名", "模板产品SPUID", "商品名称",
        "英文名称", "同规格应用", "商品规格-货号",
        "主图索引", "详情图索引", "完成情况"
    ]

    for col, title in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=title)

    sizes = ["S", "M", "L", "XL", "2XL"]

    for i, filename in enumerate(image_files, start=2):
        base_name = os.path.splitext(filename)[0].replace("-2", "")

        sheet.cell(row=i, column=1, value=os.path.abspath(image_folder))
        sheet.cell(row=i, column=2, value=filename)
        sheet.cell(row=i, column=3, value=filename)
        sheet.cell(row=i, column=4, value=filename)

        spec = "|".join([f"{base_name}-{s}" for s in sizes])
        sheet.cell(row=i, column=5, value=spec)

        sheet.cell(row=i, column=7, value=config["shop_name"])
        sheet.cell(row=i, column=8, value=config["spuid"])
        sheet.cell(row=i, column=11, value="全部")
        sheet.cell(row=i, column=12, value=base_name)
        sheet.cell(row=i, column=13, value="1")
        sheet.cell(row=i, column=14, value="1")

    wb.save(output_excel)

    messagebox.showinfo("完成", f"Excel已生成：\n{output_excel}")

# ================= UI =================
root = tk.Tk()
root.title("自动生成Excel工具（最终版）")
root.geometry("650x550")

phone_var = tk.StringVar(value=config["phone"])
account_var = tk.StringVar(value=config["account_name"])
image_var = tk.StringVar(value=config["image_dir"])
shop_var = tk.StringVar(value=config["shop_name"])
spuid_var = tk.StringVar(value=config["spuid"])
output_var = tk.StringVar(value=config.get("output_excel", ""))

def choose_image():
    path = filedialog.askdirectory()
    if path:
        image_var.set(path)

def choose_output():
    path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")],
        initialfile="上架表模板MX_已处理.xlsx"
    )
    if path:
        output_var.set(path)


tk.Label(root, text="用户名").pack()
tk.Entry(root, textvariable=account_var).pack()

tk.Label(root, text="手机号").pack()
tk.Entry(root, textvariable=phone_var).pack()

tk.Label(root, text="图片文件夹").pack()
tk.Entry(root, textvariable=image_var, width=60).pack()
tk.Button(root, text="选择文件夹", command=choose_image).pack()

tk.Label(root, text="店铺名(G列)").pack()
tk.Entry(root, textvariable=shop_var).pack()

tk.Label(root, text="SPUID(H列)").pack()
tk.Entry(root, textvariable=spuid_var).pack()

tk.Label(root, text="导出Excel路径").pack()
tk.Entry(root, textvariable=output_var, width=60).pack()
tk.Button(root, text="选择保存位置", command=choose_output).pack()

tk.Button(root, text="🚀 生成Excel", bg="green", fg="white", command=process).pack(pady=20)

root.mainloop()