# -*- coding: utf-8 -*-
import concurrent.futures
import openpyxl
import requests
import base64
import threading
import io
import json
from pathlib import Path
from PIL import Image
import tkinter as tk
from tkinter import filedialog, messagebox
import queue

# ================= 文件路径 =================
CACHE_FILE = Path("config_cache.json")

stop_event = threading.Event()
log_queue = queue.Queue()
LOCK = threading.Lock()


# ================= 默认配置 =================
DEFAULT_CONFIG = {
    "phone": "授权码",
    "account_name": "账户名",
    "api_key": "模型API密钥",
    "base_url": "https://xiaoai.plus/v1",
    "model": "gpt-4o-mini",
    "workers": 35,
    "excel_path": "",
    "prompt": """你是一个墨西哥跨境电商专家，请为这个单件T恤图片生成西班牙语商品标题：
标题要求：格式按照：超短核心特征词+商品核心关键词+特征词1+替代关键词1+特征词2+替代关键词2+特征词n+替代关键词n+适用场景/人群+尺寸/颜色列举+类目大词+热搜词；
保持语言自然地道，字符数不超过350；
不要表情和+-|/[]符号，必须为西班牙语，不要包含套装等词语，卖的是单件T恤"
输出一个自然流畅、有吸引力的西班牙语标题。
"""
}

# ================= JSON配置系统 =================
def load_config():
    if CACHE_FILE.exists():
        try:
            data = json.loads(CACHE_FILE.read_text(encoding="utf-8"))
            return {**DEFAULT_CONFIG, **data}
        except:
            return DEFAULT_CONFIG
    else:
        save_config(DEFAULT_CONFIG)
        return DEFAULT_CONFIG

def save_config(data):
    CACHE_FILE.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

config = load_config()

# ================= 日志 =================
def log(msg):
    print(msg)
    log_queue.put(msg)

# ================= 授权 =================
def check_auth(phone, account_name):
    url = "http://hk.ludada.vip:6688/check"

    data = {
        "phone": phone,
        "account_name": account_name
    }

    try:
        res = requests.post(url, json=data, timeout=10)
        result = res.json()

        if result.get("code") == 1:
            return True, result.get("msg", "OK")
        else:
            return False, result.get("msg", "未授权")

    except Exception as e:
        return False, str(e)

# ================= 图片压缩 =================
def compress_image(img_path):
    try:
        img = Image.open(img_path)

        max_size = 512
        w, h = img.size

        if w > h:
            nw = max_size
            nh = int(h * max_size / w)
        else:
            nh = max_size
            nw = int(w * max_size / h)

        img = img.resize((nw, nh), Image.LANCZOS)

        if img.mode != "RGB":
            img = img.convert("RGB")

        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=70)

        return buf.getvalue()

    except Exception as e:
        log(f"压缩失败：{e}")
        return None

# ================= API =================
def generate_title(base64_image):
    """
    成功时返回 (标题文本, None)
    失败时返回 (None, 错误详情字符串)
    """
    try:
        headers = {
            "Authorization": f"Bearer {config['api_key']}"
        }

        payload = {
            "model": config["model"],
            "messages": [{
                "role": "user",
                "content": [
                    {"type": "text", "text": config["prompt"]},
                    {"type": "image_url", "image_url": {
                        "url": f"data:image/jpeg;base64,{base64_image}"
                    }}
                ]
            }],
            "max_tokens": 250
        }

        r = requests.post(
            f"{config['base_url']}/chat/completions",
            headers=headers,
            json=payload,
            timeout=30
        )

        data = r.json()

        # API 返回了 error 字段（模型不支持/未开通/参数错误等）
        if "error" in data:
            err_msg = data["error"].get("message", str(data["error"])) if isinstance(data["error"], dict) else str(data["error"])
            return None, f"HTTP{r.status_code} | {err_msg}"

        if "choices" not in data:
            return None, f"HTTP{r.status_code} | 返回异常: {json.dumps(data, ensure_ascii=False)[:200]}"

        return data['choices'][0]['message']['content'], None

    except requests.exceptions.Timeout:
        return None, "请求超时"
    except requests.exceptions.RequestException as e:
        return None, f"网络错误: {e}"
    except Exception as e:
        return None, f"未知异常: {e}"

# ================= 单行处理 =================
def process_row(row, base_path, filename):
    try:
        img_path = (Path(base_path) / filename).resolve()

        if not img_path.exists():
            return {"row": row, "status": "fail", "title": "FILE_NOT_FOUND"}

        img = compress_image(img_path)
        if not img:
            return {"row": row, "status": "fail", "title": "COMPRESS_FAIL"}

        base64_img = base64.b64encode(img).decode()
        title, err = generate_title(base64_img)

        if not title:
            return {"row": row, "status": "fail", "title": f"API_FAIL | {err}"}

        return {"row": row, "status": "success", "title": title.strip()}

    except Exception as e:
        return {"row": row, "status": "fail", "title": str(e)}

# ================= 主任务 =================
def run_task():
    wb = openpyxl.load_workbook(config["excel_path"])
    sheet = wb.active

    total = sheet.max_row - 1
    done = 0

    with concurrent.futures.ThreadPoolExecutor(max_workers=config["workers"]) as executor:

        futures = {}

        for row in range(2, sheet.max_row + 1):

            if stop_event.is_set():
                break

            base_path = sheet.cell(row=row, column=1).value
            filename = sheet.cell(row=row, column=2).value

            # 跳过已处理
            if sheet.cell(row=row, column=9).value not in [None, "", "FAIL"]:
                log(f"[跳过] 第{row}行")
                continue

            if not base_path or not filename:
                continue

            futures[executor.submit(process_row, row, base_path, filename)] = row

        for f in concurrent.futures.as_completed(futures):

            if stop_event.is_set():
                break

            res = f.result()

            row = res["row"]
            status = res["status"]
            title = res["title"]

            if status == "success":
                with LOCK:
                    sheet.cell(row=row, column=9).value = title
                    wb.save(config["excel_path"])

                done += 1
                log(f"[成功] {done}/{total} 第{row}行 | {title[:40]}")

            else:
                with LOCK:
                    # Excel 里只写 FAIL 前缀，方便"跳过已处理"逻辑判断；详细原因看日志
                    sheet.cell(row=row, column=9).value = "FAIL"
                    wb.save(config["excel_path"])

                log(f"[失败] 第{row}行 | {title}")

    log("🎉 全部完成")

# ================= UI =================
def start():
    config["phone"] = phone_var.get()
    config["account_name"] = account_var.get()
    config["api_key"] = api_key_var.get()
    config["base_url"] = base_url_var.get()
    config["model"] = model_var.get()
    config["workers"] = int(workers_var.get())
    config["prompt"] = prompt_text.get("1.0", tk.END)

    if not config["phone"] or not config["account_name"]:
        messagebox.showerror("错误", "请填写手机号和用户名")
        return

    # 保存JSON（自动生成）
    save_config(config)
    log("💾 已自动保存JSON配置文件")

    log("🔐 正在验证授权...")

    ok, msg = check_auth(config["phone"], config["account_name"])

    if not ok:
        messagebox.showerror("授权失败", msg)
        log(f"❌ 授权失败：{msg}")
        return

    log("✅ 授权通过")

    if not config["excel_path"]:
        messagebox.showerror("错误", "请选择Excel文件")
        return

    stop_event.clear()
    threading.Thread(target=run_task, daemon=True).start()

def stop():
    stop_event.set()
    log("⛔ 已停止")

def choose_file():
    path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
    if path:
        config["excel_path"] = path
        file_label.config(text=path)

# ================= UI =================
root = tk.Tk()
root.title("服装标题生成系统")
root.geometry("900x750")

phone_var = tk.StringVar(value=config["phone"])
account_var = tk.StringVar(value=config["account_name"])
api_key_var = tk.StringVar(value=config["api_key"])
base_url_var = tk.StringVar(value=config["base_url"])
model_var = tk.StringVar(value=config["model"])
workers_var = tk.StringVar(value=str(config["workers"]))


tk.Label(root, text="账户名").pack()
tk.Entry(root, textvariable=account_var).pack()

tk.Label(root, text="授权码").pack()
tk.Entry(root, textvariable=phone_var).pack()

tk.Label(root, text="API KEY").pack()
tk.Entry(root, textvariable=api_key_var, width=90).pack()

tk.Label(root, text="BASE URL").pack()
tk.Entry(root, textvariable=base_url_var, width=90).pack()

tk.Label(root, text="MODEL").pack()
tk.Entry(root, textvariable=model_var).pack()

tk.Label(root, text="线程速（不超35）").pack()
tk.Entry(root, textvariable=workers_var).pack()

tk.Button(root, text="选择Excel", command=choose_file).pack()
file_label = tk.Label(root, text=config["excel_path"])
file_label.pack()

tk.Label(root, text="提示词").pack()
prompt_text = tk.Text(root, height=12)
prompt_text.insert(tk.END, config["prompt"])
prompt_text.pack()

tk.Button(root, text="🚀 开始", bg="green", fg="white", command=start).pack()
tk.Button(root, text="⛔ 停止", bg="red", fg="white", command=stop).pack()

log_box = tk.Text(root, height=18)
log_box.pack(fill="both", expand=True)

def update_log():
    while not log_queue.empty():
        log_box.insert(tk.END, log_queue.get() + "\n")
        log_box.see(tk.END)
    root.after(200, update_log)

root.after(200, update_log)
root.mainloop()