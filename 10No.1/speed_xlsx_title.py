# -*- coding: utf-8 -*-
import concurrent.futures
import openpyxl
import requests
import base64
import logging
import threading
import io
from openpyxl import load_workbook
from pathlib import Path
from requests.exceptions import Timeout
from PIL import Image  # 用于图片压缩

# 配置日志
logging.basicConfig(
    filename=str(Path(__file__).parent / 'title_generation.log'),
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    encoding='utf-8'
)
print(f"📝 日志文件位置：{Path(__file__).parent / 'title_generation.log'}")

# 全局配置
CURRENT_DIR = Path(__file__).parent
EXCEL_PATH = CURRENT_DIR / "上架表模板MX_已处理.xlsx"
API_KEY = "sk-4BcnxDna9rNMGe5XqRnOHNlQdCEuZhhCvb2LwWUJR7f5iYaG"
BASE_URL = "https://xiaoai.plus/v1"
TIMEOUT = 35
MAX_WORKERS = 35  # 降低线程数以避免API限制
LOCK = threading.Lock()  # 线程锁

def compress_image(img_path):
    """压缩图片以节省token"""
    try:
        # 打开图片并调整大小
        img = Image.open(img_path)
        
        # 计算新尺寸（最大边长为512像素）
        max_size = 512
        width, height = img.size
        if width > height:
            new_width = max_size
            new_height = int(height * max_size / width)
        else:
            new_height = max_size
            new_width = int(width * max_size / height)
        
        # 调整尺寸并转换格式
        img = img.resize((new_width, new_height), Image.LANCZOS)
        
        # 转换为JPEG格式并设置质量
        if img.mode != 'RGB':
            img = img.convert('RGB')
        
        # 保存到内存字节流
        buffer = io.BytesIO()
        img.save(buffer, format="JPEG", quality=70)  # 中等质量
        
        return buffer.getvalue()
    except Exception as e:
        logging.error(f"图片压缩失败：{img_path} - {str(e)}")
        return None

def process_row(sheet, row):
    """处理单行任务的独立函数"""
    base_path = sheet.cell(row=row, column=1).value
    filename = sheet.cell(row=row, column=2).value
    
    # 跳过无效行和已处理项
    if not all([base_path, filename]) or sheet.cell(row=row, column=9).value:
        return None
    
    img_path = (CURRENT_DIR / base_path.strip() / filename.strip()).resolve()
    if not img_path.exists():
        logging.warning(f"Missing file: {img_path}")
        return (row, None)
    
    try:
        # 压缩图片
        compressed_img = compress_image(img_path)
        if compressed_img is None:
            logging.error(f"图片压缩失败，使用原始图片：{img_path}")
            compressed_img = img_path.read_bytes()
        
        base64_image = base64.b64encode(compressed_img).decode('utf-8')
        title = generate_title(base64_image)
        return (row, title.strip('"') if title else None)
    except Exception as e:
        logging.error(f"行处理失败：第{row}行 - {str(e)}")
        return (row, None)

def generate_title(image_base64):
    """生成标题的函数，优化提示词以节省token"""
    headers = {"Authorization": f"Bearer {API_KEY}"}
    
    # 优化提示词：更简洁，节省token
    prompt_text = (
"你是一个墨西哥跨境电商专家，请为这个单件T恤图片生成西班牙语商品标题："
"标题要求："
"【超短核心特征词+商品核心关键词】+【特征词1+替代关键词1】+【特征词2+替代关键词2】 .… 【特征词n+ 替代关键词n+适用场景/人群】+ 【尺寸/颜色列举】+ 【类目大词】+【热搜词】"
"保持语言自然地道，字符数不超过350；"
"不要有表情等符号，必须为西班牙语，不要包含套装等词语，卖的是单件T恤"
"输出一个自然流畅、有吸引力的西班牙语标题。"

    )
    
    payload = {
        "model": "gpt-4o",
        "messages": [{
            "role": "user",
            "content": [
                {"type": "text", "text": prompt_text},
                {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{image_base64}"}}
            ]
        }],
        "max_tokens": 250  # 限制最大token数
    }
    
    try:
        response = requests.post(
            f"{BASE_URL}/chat/completions", 
            headers=headers, 
            json=payload, 
            timeout=TIMEOUT
        )
        response.raise_for_status()
        return response.json()['choices'][0]['message']['content']
    except Exception as e:
        logging.error(f"API请求失败：{str(e)}")
        return None

def main():
    print("🟢 多线程程序启动")
    if not EXCEL_PATH.exists():
        print(f"❌ 错误：文件 {EXCEL_PATH} 不存在")
        return

    wb = load_workbook(EXCEL_PATH)
    sheet = wb.active
    total_rows = sheet.max_row - 1  # 排除标题行
    processed_count = 0
    
    # 创建线程池
    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        # 收集所有任务
        tasks = {}
        for row in range(2, sheet.max_row + 1):
            # 跳过已处理的行
            if sheet.cell(row=row, column=9).value:
                continue
                
            tasks[executor.submit(process_row, sheet, row)] = row
        
        # 处理结果
        for future in concurrent.futures.as_completed(tasks):
            result = future.result()
            row = tasks[future]
            
            if result is None:
                print(f"⏩ 跳过第{row}行")
                continue
                
            row_num, title = result
            if title:
                # 线程安全写入
                with LOCK:
                    sheet.cell(row=row_num, column=9).value = title
                    wb.save(EXCEL_PATH)  # 每次保存都写盘
                
                processed_count += 1
                print(f"✅ 进度：{processed_count}/{total_rows} | 第{row_num}行：{title[:30]}...")
            else:
                print(f"⚠️ 第{row_num}行处理失败")

    print(f"\n🎉 处理完成！共处理 {processed_count}/{total_rows} 行")

if __name__ == "__main__":
    main()