# -*- coding: utf-8 -*-
import openpyxl
import requests
import base64
import logging
from openpyxl import load_workbook
from pathlib import Path
from requests.exceptions import Timeout

# 配置日志
logging.basicConfig(
    filename='xlsx_title.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    encoding='utf-8'
)

# 自动获取同目录文件路径
CURRENT_DIR = Path(__file__).parent
EXCEL_PATH = CURRENT_DIR / "上架表模板MX.xlsx"  # 自动定位同目录文件
API_KEY = "sk-4BcnxDna9rNMGe5XqRnOHNlQdCEuZhhCvb2LwWUJR7f5iYaG"  # 替换为有效API密钥
BASE_URL = "https://xiaoai.plus/v1"
TIMEOUT = 10  # API请求超时时间（秒）

def encode_image(image_path):
    """处理图片并返回base64编码"""
    try:
        logging.info(f"Processing image: {image_path}")
        with open(image_path, "rb") as image_file:
            return base64.b64encode(image_file.read()).decode('utf-8')
    except Exception as e:
        logging.error(f"图片处理失败：{image_path} - {str(e)}")
        return None

def generate_title(image_base64):
    """调用GPT-4o生成标题"""
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {API_KEY}"
    }

    payload = {
        "model": "gpt-4o",
        "messages": [
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": (
                            "作为墨西哥电商专家，为此T恤生成西班牙语商品标题。要求："
                            "1. 包含墨西哥热搜关键词 2. 突出设计图案亮点 3. 符合当地用语"
                            "4. 长度≤80字符 5. 不要使用引号 6.规避违规词语 7.符合潮流"
                        )
                    },
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/jpeg;base64,{image_base64}"
                        }
                    }
                ]
            }
        ],
        "max_tokens": 300
    }

    try:
        response = requests.post(
            f"{BASE_URL}/chat/completions",
            headers=headers,
            json=payload,
            timeout=TIMEOUT
        )
        response.raise_for_status()
        return response.json()['choices'][0]['message']['content']
    except Timeout:
        logging.error("API请求超时")
        return None
    except Exception as e:
        logging.error(f"API请求失败：{str(e)}")
        return None

def main():
    try:
        print("🟢 程序启动")
        wb = load_workbook(EXCEL_PATH)
        sheet = wb.active

        for row in range(2, sheet.max_row + 1):
            # 关键修改：从A列获取基础路径，B列获取文件名
            base_path = sheet.cell(row=row, column=1).value  # A列
            filename = sheet.cell(row=row, column=2).value    # B列
            
            if not all([base_path, filename]):
                print(f"⏩ 跳过第{row}行：缺少路径或文件名")
                continue

            if sheet.cell(row=row, column=9).value:  # 跳过已处理项
                continue

            print(f"\n🔍 处理第{row}行")
            img_path = Path(str(base_path).strip()) / str(filename).strip()
            print(f"📁 基础路径：{base_path}")
            print(f"📷 文件名：{filename}")
            print(f"🛣️ 完整路径：{img_path}")

            if not img_path.exists():
                print(f"❌ 文件不存在，请检查：{img_path}")
                logging.warning(f"Missing file: {img_path}")
                continue

            if base64_image := encode_image(img_path):
                print("🔄 正在生成标题...")
                if title := generate_title(base64_image):
                    sheet.cell(row=row, column=9).value = title.strip('"')
                    print(f"✅ 生成成功：{title[:45]}...")
                    wb.save(EXCEL_PATH)  # 逐行保存
                else:
                    print("⚠️ 标题生成失败")
            else:
                print("⚠️ 图片处理失败")

        print("\n🎉 处理完成！请检查I列结果")
    except Exception as e:
        logging.critical(f"主流程错误：{str(e)}")
        print(f"❌ 发生严重错误：{str(e)}")

if __name__ == "__main__":
    main()