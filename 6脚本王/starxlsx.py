import os
from openpyxl import load_workbook, Workbook

def write_image_names_to_excel(image_folder, source_excel, output_suffix=""):
#def write_image_names_to_excel(image_folder, source_excel, output_suffix="_已处理"):
    """
    将图片信息写入Excel（保留源文件不变）：
    - 生成新文件：原文件名 + "_已处理.xlsx"
    - 主要功能：
      1. 将S列(第19列)商品名称复制到I列(第9列)
      2. 删除O列(第15列)及之后的所有列
      3. 其他原有功能保持不变
    """
    try:
        # 验证图片文件夹
        if not os.path.exists(image_folder):
            raise FileNotFoundError(f"图片文件夹不存在: {os.path.abspath(image_folder)}")
        
        if not os.path.isdir(image_folder):
            raise NotADirectoryError(f"路径不是文件夹: {os.path.abspath(image_folder)}")

        # 获取并处理图片文件名（按名称排序）
        image_files = sorted([f for f in os.listdir(image_folder) 
                          if f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp'))])
        
        if not image_files:
            print(f"⚠️ 文件夹 {os.path.abspath(image_folder)} 中未找到图片文件！")
            return

        print(f"🔍 找到 {len(image_files)} 个图片文件")

        # 设置输出文件名
        file_name, file_ext = os.path.splitext(source_excel)
        output_excel = f"{file_name}{output_suffix}{file_ext}"
        
        # 加载Excel文件
        if os.path.exists(source_excel):
            wb = load_workbook(source_excel)
            sheet = wb.active
            print(f"📂 已加载源文件: {os.path.abspath(source_excel)}")
            
            # 删除O列及之后的所有列（如果存在）
            #if sheet.max_column >= 15:
            #    sheet.delete_cols(15, sheet.max_column - 14)
            #    print(f"🗑️ 已删除O列及之后的{sheet.max_column - 14}列")
        else:
            wb = Workbook()
            sheet = wb.active
            print("⚠️ 源文件不存在，创建新文件")
            # 添加表头
            headers = {
                1: "素材本机地址",
                2: "素材图",
                3: "详情图list", 
                4: "主图list",
                5: "SKU货号编码list",
                6: "本地图片包地址下文件列表",
                7: "店铺名",
                8: "模板产品SPUID",
                9: "商品名称",
                10: "英文名称",
                11: "同规格应用",
                12: "商品规格-货号(选填)",
                13: "主图索引",
                14: "详情图索引"
            }
            for col, title in headers.items():
                sheet.cell(row=1, column=col, value=title)

        # 处理文件名（去掉后缀和"-2"）
        base_names = []
        for filename in image_files:
            name = os.path.splitext(filename)[0]  # 去掉后缀
            name = name.replace("-2", "")         # 去掉"-2"
            #if name.startswith("LU") and len(name) >= 5:
            #    name = f"LU{name[3:]}"  # 去掉第一个0
            base_names.append(name)

        # 规格组合配置
        sizes = ["S", "M", "L", "XL", "2XL"]
        
        # 固定值配置
        fixed_values = {
            7: "老六本土",            # G列：店铺名
            8: "602908296853942",    # H列：商品规格-货号
            11: "全部",              # K列：同规格应用
            13: "1",                 # M列：主图索引
            14: "1"                  # N列：详情图索引
        }
        
        # 从第2行开始写入数据
        start_row = 2
        for i, (filename, base_name) in enumerate(zip(image_files, base_names), start=start_row):
            # A列：完整文件夹路径
            sheet.cell(row=i, column=1, value=os.path.abspath(image_folder))
            
            # B/C/D列：完整文件名
            for col in [2, 3, 4]:
                sheet.cell(row=i, column=col, value=filename)
            
            # E列：规格组合
            spec_combinations = "|".join([f"{base_name}-B-{size}" for size in sizes])
            sheet.cell(row=i, column=5, value=spec_combinations)
            
            # L列：店铺名（纯净文件名）
            sheet.cell(row=i, column=12, value=base_name)
            
            # 复制商品名称(S列)到I列（如果S列存在）
            if sheet.max_column >= 19 and sheet.cell(row=i, column=19).value:
                sheet.cell(row=i, column=9).value = sheet.cell(row=i, column=19).value
            
            # 写入各固定值列
            for col, val in fixed_values.items():
                sheet.cell(row=i, column=col, value=val)

        # 再次确保删除O列及之后的列（防止新增列）
        if sheet.max_column >= 15:
            sheet.delete_cols(15, sheet.max_column - 14)

        # 保存新文件
        wb.save(output_excel)
        print(f"\n✅ 处理完成：\n"
              f"- 源文件保留: {os.path.abspath(source_excel)}\n"
              f"- 生成新文件: {os.path.abspath(output_excel)}\n"
              f"- 已处理记录: {len(image_files)} 条\n"
              f"- 商品名称已从S列复制到I列\n"
              f"- 已删除O列及之后的所有列")
        
    except Exception as e:
        print(f"\n❌ 发生错误: {str(e)}")

if __name__ == "__main__":
    # 配置路径（自动获取当前脚本所在目录）
    current_dir = os.path.dirname(os.path.abspath(__file__))
    image_dir = os.path.join(current_dir, "images")
    source_excel = os.path.join(current_dir, "上架表模板MX.xlsx")
    
    if not os.path.exists(image_dir):
        print(f"⚠️ 创建图片文件夹: {image_dir}")
        os.makedirs(image_dir)
        print("请将图片放入文件夹后重新运行")
    else:
        write_image_names_to_excel(image_dir, source_excel)