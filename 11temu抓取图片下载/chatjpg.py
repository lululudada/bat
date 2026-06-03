import os
import re
import time
import requests
from openpyxl import load_workbook
import concurrent.futures
from tqdm import tqdm
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# ========== 配置参数（请根据需要修改）==========
EXCEL_FILE = "data.xlsx"            # Excel 文件名
SAVE_FOLDER = "downloaded_images"   # 保存图片的文件夹
URL_COLUMN = "D"                    # 存放图片 URL 的列
NAME_COLUMN = "L"                   # 存放自定义文件名的列
START_ROW = 2                       # 起始行（第一行为表头）
MAX_THREADS = 100                    # 最大线程数（建议 5~15）
USE_HTTP_FALLBACK = True            # 是否将 HTTPS 替换为 HTTP（测试用，生产环境建议 False）
RETRY_TIMES = 1                     # 下载失败重试次数
CONNECT_TIMEOUT = 5                 # 连接超时（秒）
READ_TIMEOUT = 30                   # 读取超时（秒）
# ============================================

def sanitize_filename(name):
    """移除文件名中的非法字符"""
    return re.sub(r'[\\/*?:"<>|]', '_', name)

def get_extension_from_url(url):
    """从 URL 中提取文件扩展名（含点号），如 '.jpg'，找不到则返回空字符串"""
    basename = url.split('/')[-1].split('?')[0]
    if '.' in basename:
        ext = '.' + basename.split('.')[-1]
        if ext.lower() in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.tiff']:
            return ext
    return ''

def download_image(url, save_path, retry_times=RETRY_TIMES):
    """下载图片，支持重试和指数退避"""
    session = requests.Session()
    # 配置重试策略（针对连接错误和超时）
    retry_strategy = Retry(
        total=retry_times,
        backoff_factor=1,  # 退避因子：1, 2, 4, 8...
        status_forcelist=[500, 502, 503, 504],
        allowed_methods=["GET"]
    )
    adapter = HTTPAdapter(max_retries=retry_strategy)
    session.mount("http://", adapter)
    session.mount("https://", adapter)

    for attempt in range(1, retry_times + 1):
        try:
            response = session.get(
                url,
                stream=True,
                timeout=(CONNECT_TIMEOUT, READ_TIMEOUT)
            )
            if response.status_code == 200:
                # 确保目录存在
                os.makedirs(os.path.dirname(save_path), exist_ok=True)
                with open(save_path, 'wb') as f:
                    for chunk in response.iter_content(1024):
                        f.write(chunk)
                return True, None
            else:
                return False, f"HTTP {response.status_code}"
        except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as e:
            error_msg = f"尝试 {attempt}/{retry_times} 失败: {str(e)}"
            if attempt == retry_times:
                return False, error_msg
            # 指数退避等待
            wait_time = 2 ** (attempt - 1)
            tqdm.write(f"下载超时，{wait_time}秒后进行第 {attempt+1} 次重试...")
            time.sleep(wait_time)
        except Exception as e:
            return False, str(e)
    return False, "未知错误"

def download_task(task):
    """处理单个下载任务"""
    url, save_path = task
    success, error = download_image(url, save_path)
    if success:
        tqdm.write(f"✓ 下载成功：{os.path.basename(save_path)}")
    else:
        tqdm.write(f"✗ 下载失败：{url} - 错误：{error}")
    return success

# 创建保存目录
os.makedirs(SAVE_FOLDER, exist_ok=True)

# 加载 Excel 文件
wb = load_workbook(EXCEL_FILE)
ws = wb.active

# 获取列索引
url_col_idx = ord(URL_COLUMN.upper()) - 64
name_col_idx = ord(NAME_COLUMN.upper()) - 64

# 收集下载任务
tasks = []
existing_files = set(os.listdir(SAVE_FOLDER))

# 计算需要遍历的最小列和最大列，避免无用单元格
min_col = min(url_col_idx, name_col_idx)
max_col = max(url_col_idx, name_col_idx)

for row in ws.iter_rows(min_row=START_ROW, min_col=min_col, max_col=max_col):
    # 找到 URL 和 NAME 列在当前行中的相对位置
    url_cell = row[url_col_idx - min_col]
    name_cell = row[name_col_idx - min_col] if name_col_idx >= min_col else None

    url_value = url_cell.value
    if not url_value:
        continue

    # 处理多行 URL（只取第一个有效 URL）
    first_url = str(url_value).strip().split('\n')[0].strip()
    if not first_url:
        continue

    # 可选：将 HTTPS 替换为 HTTP（测试用）
    if USE_HTTP_FALLBACK and first_url.startswith('https://'):
        first_url = 'http://' + first_url[8:]

    # 确定文件名
    custom_name = name_cell.value.strip() if name_cell and name_cell.value else ''
    if custom_name:
        if '.' in custom_name:
            # 用户已包含扩展名
            filename = sanitize_filename(custom_name)
        else:
            ext = get_extension_from_url(first_url)
            if not ext:
                tqdm.write(f"警告：无法从 URL 获取扩展名，跳过 {first_url}")
                continue
            filename = sanitize_filename(custom_name) + ext
    else:
        # 回退：使用 URL 自带文件名
        filename = os.path.basename(first_url.split('/')[-1].split('?')[0])
        if not filename or '.' not in filename:
            tqdm.write(f"警告：无法从 URL 提取有效文件名，跳过 {first_url}")
            continue

    save_path = os.path.join(SAVE_FOLDER, filename)

    # 跳过已存在的文件
    if filename in existing_files:
        tqdm.write(f"文件已存在，跳过：{filename}")
        continue

    existing_files.add(filename)
    tasks.append((first_url, save_path))

# 执行多线程下载
if tasks:
    print(f"共需下载 {len(tasks)} 个文件")
    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_THREADS) as executor:
        futures = [executor.submit(download_task, task) for task in tasks]
        with tqdm(total=len(futures), desc="下载进度", unit="file") as pbar:
            for future in concurrent.futures.as_completed(futures):
                future.result()
                pbar.update(1)
else:
    print("没有需要下载的文件")

print("全部任务处理完成！")