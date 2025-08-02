import os
import glob
import shutil
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def find_template_file():
    """在当前目录中查找包含'Bulk_Upload_Template'的Excel文件"""
    possible_files = glob.glob('*Bulk_Upload_Template*.xlsx')
    if not possible_files:
        raise FileNotFoundError("未找到包含'Bulk_Upload_Template'的Excel文件")
    return possible_files[0]

def process_bulk_upload_template():
    # 查找模板文件
    input_file_path = find_template_file()
    print(f"找到模板文件: {input_file_path}")
    
    # 读取Excel文件信息（不加载实际数据）
    print("读取文件结构...")
    wb = load_workbook(input_file_path, read_only=True)
    
    # 获取Template工作表
    if 'Template' not in wb.sheetnames:
        raise ValueError("Excel文件中缺少'Template'工作表")
    
    # 获取Template工作表的前两行
    template_sheet = wb['Template']
    header_rows = []
    for i, row in enumerate(template_sheet.iter_rows(values_only=True)):
        if i < 2:
            header_rows.append(row)
        else:
            break
    
    # 关闭只读工作簿
    wb.close()
    
    # 使用pandas高效读取数据
    print("使用pandas读取数据...")
    start_time = time.time()
    template_df = pd.read_excel(input_file_path, sheet_name='Template', engine='openpyxl')
    print(f"数据读取完成 (耗时: {time.time()-start_time:.2f}秒)")
    
    # 按Goods ID分组
    print("分组处理数据...")
    grouped = template_df.groupby('Goods ID')
    total_products = len(grouped)
    print(f"总货品数量: {total_products}")
    
    # 计算需要拆分的文件数量
    num_files = (total_products + 1998) // 1999
    print(f"需要拆分成 {num_files} 个文件")
    
    # 获取分组键（Goods ID）
    product_ids = list(grouped.groups.keys())
    
    # 处理每个文件
    for file_num in range(1, num_files + 1):
        file_start_time = time.time()
        print(f"\n开始处理文件 #{file_num}...")
        
        # 创建新文件名
        output_path = os.path.splitext(input_file_path)[0] + f"_已处理_{file_num}.xlsx"
        
        # 创建新工作簿（不复制原始文件）
        output_wb = load_workbook(input_file_path)
        
        # 获取Template工作表
        if 'Template' in output_wb.sheetnames:
            output_template = output_wb['Template']
        else:
            output_template = output_wb.create_sheet('Template')
        
        # 计算当前文件的货品范围
        start_idx = (file_num - 1) * 1999
        end_idx = min(file_num * 1999, total_products)
        current_product_ids = product_ids[start_idx:end_idx]
        
        # 获取当前批次的数据
        current_df = pd.concat([grouped.get_group(pid) for pid in current_product_ids])
        
        # 高效清除数据（保留前两行）
        print("准备数据区域...")
        # 高效方法：直接覆盖前两行，然后写入新数据
        for r_idx, row in enumerate(header_rows, 1):
            for c_idx, value in enumerate(row, 1):
                output_template.cell(row=r_idx, column=c_idx, value=value)
        
        # 清除第三行及之后的数据
        if output_template.max_row > 2:
            output_template.delete_rows(3, output_template.max_row - 2)
        
        # 添加新数据行
        print(f"添加新数据 ({len(current_df)} 行)...")
        for r_idx, row in enumerate(dataframe_to_rows(current_df, index=False, header=False), 3):
            output_template.append(row)
        
        # 保存修改
        print("保存文件...")
        output_wb.save(output_path)
        
        file_time = time.time() - file_start_time
        print(f"文件 #{file_num} 完成! (耗时: {file_time:.2f}秒) 包含货品 {start_idx+1} 到 {end_idx}")
    
    total_time = time.time() - start_time
    return f"处理完成! 共创建了 {num_files} 个文件 (总耗时: {total_time:.2f}秒)"

# 使用示例
if __name__ == "__main__":
    try:
        print("开始处理批量上传模板...")
        result = process_bulk_upload_template()
        print("\n" + result)
        print("按Enter键退出...")
        input()
    except Exception as e:
        print(f"处理过程中出错: {str(e)}")
        print("按Enter键退出...")
        input()