# -*- coding: utf-8 -*-
import openpyxl
import requests
import base64
import logging
import re
from openpyxl import Workbook, load_workbook  # 添加缺失的导入
from pathlib import Path
from requests.exceptions import Timeout

# 获取当前脚本所在目录
CURRENT_DIR = Path(__file__).parent

# 配置日志文件路径（与脚本同目录）
LOG_PATH = CURRENT_DIR / 'file_title.log'  # 新增日志路径定义
EXCEL_PATH = CURRENT_DIR / "outdoor_products.xlsx"
IMAGE_DIR = CURRENT_DIR / "images"

# 日志配置（新增文件模式参数）
logging.basicConfig(
    filename=str(LOG_PATH),  # 转换为字符串路径
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    encoding='utf-8',
    filemode='a'  # 追加写入模式（默认值）
)


# 配置参数
CURRENT_DIR = Path(__file__).parent
EXCEL_PATH = CURRENT_DIR / "outdoor_products.xlsx"
IMAGE_DIR = CURRENT_DIR / "images"  # 图片目录
API_KEY = "sk-4BcnxDna9rNMGe5XqRnOHNlQdCEuZhhCvb2LwWUJR7f5iYaG"
BASE_URL = "https://xiaoai.plus/v1"
TIMEOUT = 10  # API请求超时时间（秒）
SUPPORTED_EXT = ('.jpg', '.jpeg', '.png', '.webp')  # 支持的图片格式

def init_excel():
    """创建或初始化Excel文件"""
    if not EXCEL_PATH.exists():
        wb = Workbook()
        ws = wb.active
        ws.append(["路径", "文件名", "商品标题", "详情介绍", "处理状态"])
        wb.save(EXCEL_PATH)
        logging.info("新建Excel文件成功")

def scan_images():
    """扫描图片目录并返回结构化的图片数据"""
    if not IMAGE_DIR.exists():
        raise FileNotFoundError(f"图片目录不存在：{IMAGE_DIR}")

    image_data = []
    for img_path in IMAGE_DIR.glob('**/*'):
        if img_path.suffix.lower() in SUPPORTED_EXT and img_path.is_file():
            image_data.append({
                "path": str(img_path.parent),
                "filename": img_path.name,
                "full_path": img_path
            })
    
    if not image_data:
        raise ValueError(f"未找到支持的图片文件，支持格式：{SUPPORTED_EXT}")
    
    return image_data

def update_excel(image_list):
    """更新Excel文件结构"""
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    
    # 获取现有记录
    existing = set()
    for row in range(2, ws.max_row + 1):
        path = ws.cell(row, 1).value
        name = ws.cell(row, 2).value
        if path and name:
            existing.add(f"{path}/{name}")
    
    # 添加新记录
    new_items = 0
    for img in image_list:
        unique_key = f"{img['path']}/{img['filename']}"
        if unique_key not in existing:
            ws.append([
                img['path'],
                img['filename'],
                None,  # 标题占位
                None,  # 详情占位
                "PENDING"  # 初始状态
            ])
            new_items += 1
    
    if new_items > 0:
        wb.save(EXCEL_PATH)
        logging.info(f"新增{new_items}条图片记录")
    
    return wb

def validate_response(text):
    """验证API响应格式"""
    pattern = r"TÍTULO:(.*?)DESCRIPCIÓN:(.*)"
    match = re.search(pattern, text, re.DOTALL)
    return (match.group(1).strip(), match.group(2).strip()) if match else (None, None)

def encode_image(image_path):
    """处理图片并返回base64编码"""
    try:
        with open(image_path, "rb") as image_file:
            return base64.b64encode(image_file.read()).decode('utf-8')
    except Exception as e:
        logging.error(f"图片处理失败：{image_path} - {str(e)}")
        return None

def generate_outdoor_content(image_base64):
    """调用GPT-4o生成户外用品内容"""
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {API_KEY}"
    }

    prompt = """Como experto en comercio electrónico mexicano, genera para este producto outdoor:
1. TÍTULO en español con:
   - Palabras clave populares en México
   - Ventajas del diseño
   - Lenguaje coloquial local
   - Máx 80 caracteres
   - Prohibido caracteres especiales

2. DESCRIPCIÓN con:
   - 3 ventajas principales
   - Materiales duraderos
   - Escenarios de uso típicos
   - Máx 200 caracteres

Formato requerido:
TÍTULO: [texto]
DESCRIPCIÓN: [texto]"""

    payload = {
        "model": "gpt-4o",
        "messages": [
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": prompt
                    },
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/jpeg;base64,{image_base64}"
                        }
                    }
                ]
            }
        ],
        "max_tokens": 500,
        "temperature": 0.7
    }

    try:
        response = requests.post(
            f"{BASE_URL}/chat/completions",
            headers=headers,
            json=payload,
            timeout=TIMEOUT
        )
        response.raise_for_status()
        return validate_response(response.json()['choices'][0]['message']['content'])
    except Timeout:
        logging.error("API请求超时")
        return None, None
    except Exception as e:
        logging.error(f"API请求失败：{str(e)}")
        return None, None

def main():
    try:
        print("🟢 启动智能商品信息生成系统")
        
        # 初始化Excel文件
        init_excel()
        
        # 扫描图片目录
        try:
            image_list = scan_images()
            print(f"📁 发现{len(image_list)}张待处理图片")
        except Exception as e:
            print(f"❌ 错误：{str(e)}")
            return

        # 更新Excel结构
        wb = update_excel(image_list)
        sheet = wb.active

        # 处理待生成内容
        processed = 0
        for row in range(2, sheet.max_row + 1):
            status = sheet.cell(row, 5).value
            if status == "COMPLETED":
                continue

            img_path = Path(sheet.cell(row, 1).value) / sheet.cell(row, 2).value
            if not img_path.exists():
                sheet.cell(row, 5).value = "FILE_MISSING"
                continue

            if base64_image := encode_image(img_path):
                title, desc = generate_outdoor_content(base64_image)
                if title and desc:
                    sheet.cell(row, 3).value = title
                    sheet.cell(row, 4).value = desc
                    sheet.cell(row, 5).value = "COMPLETED"
                    processed += 1
                    print(f"✅ 完成：{title[:35]}...")
                else:
                    sheet.cell(row, 5).value = "API_ERROR"
                wb.save(EXCEL_PATH)  # 实时保存

        print(f"\n🎉 处理完成！共生成{processed}条商品信息")
        print(f"📊 结果文件：{EXCEL_PATH}")

    except Exception as e:
        logging.critical(f"系统错误：{str(e)}")
        print(f"🔥 发生严重错误：{str(e)}")

if __name__ == "__main__":
    main()