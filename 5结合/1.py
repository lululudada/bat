import os
from openpyxl import load_workbook, Workbook

def write_image_names_to_excel(image_folder, source_excel, output_suffix="_已处理"):
    """
    将图片信息写入Excel（保留源文件不变）：
    - 生成新文件：原文件名 + "_已处理.xlsx"
    - A列：完整图片路径（从第2行开始）
    - B列：素材图 | C列：详情图list | D列：主图list
    """
    try:
        # 验证图片文件夹是否存在
        if not os.path.exists(image_folder):
            raise FileNotFoundError(f"图片文件夹不存在: {os.path.abspath(image_folder)}")
        
        if not os.path.isdir(image_folder):
            raise NotADirectoryError(f"路径不是文件夹: {os.path.abspath(image_folder)}")

        # 获取所有图片文件（按名称排序）
        image_files = sorted([f for f in os.listdir(image_folder) 
                          if f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp'))])
        
        if not image_files:
            print(f"⚠️ 文件夹 {os.path.abspath(image_folder)} 中未找到图片文件！")
            return

        print(f"🔍 在 {os.path.abspath(image_folder)} 中找到 {len(image_files)} 个图片文件")

        # 设置输出文件名
        file_name, file_ext = os.path.splitext(source_excel)
        output_excel = f"{file_name}{output_suffix}{file_ext}"
        
        # 复制源Excel文件（保持格式不变）
        if os.path.exists(source_excel):
            wb = load_workbook(source_excel)
            sheet = wb.active
            print(f"📂 已加载源文件: {os.path.abspath(source_excel)}")
        else:
            wb = Workbook()
            sheet = wb.active
            print(f"⚠️ 源文件不存在，创建新文件")
            # 添加表头
            sheet.cell(row=1, column=1, value="素材本机地址")
            sheet.cell(row=1, column=2, value="素材图")
            sheet.cell(row=1, column=3, value="详情图list")
            sheet.cell(row=1, column=4, value="主图list")

        # 从第2行开始写入数据
        start_row = 2
        for i, filename in enumerate(image_files, start=start_row):
            # 第一列写入完整路径
            full_path = os.path.abspath(os.path.join(image_folder))
            sheet.cell(row=i, column=1, value=full_path)
            
            # 其他列写入文件名
            for col in [2, 3, 4]:  # 固定写入B/C/D列
                sheet.cell(row=i, column=col, value=filename)

        # 保存为新文件
        wb.save(output_excel)
        print(f"\n✅ 数据处理完成：\n"
              f"- 源文件保留不变: {os.path.abspath(source_excel)}\n"
              f"- 生成新文件: {os.path.abspath(output_excel)}\n"
              f"- 写入数据: {len(image_files)} 条图片记录")
        
    except Exception as e:
        print(f"\n❌ 发生错误: {str(e)}")
        return

# 使用示例
if __name__ == "__main__":
    # 配置路径（使用绝对路径更可靠）
    current_dir = os.path.dirname(os.path.abspath(__file__))
    image_dir = os.path.join(current_dir, "images")  # 图片文件夹
    source_excel = os.path.join(current_dir, "上架表模板MX.xlsx")  # 源Excel文件
    
    # 检查图片文件夹
    if not os.path.exists(image_dir):
        print(f"⚠️ 图片文件夹不存在，正在创建: {image_dir}")
        os.makedirs(image_dir)
        print(f"请将图片放入文件夹后重新运行")
    else:
        write_image_names_to_excel(image_dir, source_excel)